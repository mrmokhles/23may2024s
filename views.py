from tkinter import Image
from urllib import response
from django.shortcuts import render,HttpResponse,redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import login_required
from django.core import serializers

from app1.models import contactEnquiry
from app1.forms import EmployeeForm,AdminEmployeeForm
from app1 import views
from .filters import OrderFilters

from django.core.paginator import Paginator,EmptyPage,PageNotAnInteger
from openpyxl import Workbook, workbook
from openpyxl.styles import *
import decimal
from app1.models import CountryGDP,Image
from tablib import Dataset
from .resources import PersonResourc
from .forms import ImageForm
from .models import Person, City
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.views.generic import View
from .forms import PersonCreationForm
from .models import Person, City,EmployeeLogin

from django.http import JsonResponse
import json
from .models import Profile,Admin_add_employee,Admin_add_overtime
from .pdfprocess import html_to_pdf

from django.db.models import Q

#for reportlab
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table,TableStyle,SimpleDocTemplate,Frame, Paragraph, Spacer,Image
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.colors import pink, black, red, blue, green
from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle







# Create your views here.
@login_required(login_url='login')
def HomePage(request):
    contactdata=contactEnquiry.objects.all()
    
    filters=OrderFilters(request.GET,queryset=contactdata)
    data={'contactdata':contactdata,'filters':filters}
    return render (request,'home.html',data)

@login_required(login_url='login')
def SignupPage(request):
    if request.method=='POST':
        uname=request.POST.get('username')
        email=request.POST.get('email')
        pass1=request.POST.get('password1')
        pass2=request.POST.get('password2')

        if pass1!=pass2:
            return HttpResponse("Your password and confrom password are not Same!!")
        else:
            #my_user=User.objects.create_user(uname,email,pass1)
            my_user=EmployeeLogin(user_name=uname,user_email=email,user_password=pass1)
            my_user.save()
            return redirect('login')

    return render (request,'signup.html')

#@login_required(login_url='login')
def LoginPage(request):
    if request.method=='POST':
        username=request.POST.get('username')
        pass1=request.POST.get('pass')

        user=EmployeeLogin.objects.all()

        # for m in user:

        #     if m.user_name==username and m.user_password==pass1:
               
        #         return redirect('home')
        #     else:
        #         return HttpResponse ("Username or Password is incorrect!!!") 
            
        user=authenticate(request,username=username,password=pass1)
        if user is not None:
            login(request,user)
            return redirect('home')
        else:
            return HttpResponse ("Username or Password is incorrect!!!")

    return render (request,'login.html')

def LogoutPage(request):
    logout(request)
    return redirect('login')

@login_required(login_url='login')
def saveEnquiry(request):
    if request.method=='POST':
        name=request.POST.get('name')
        email=request.POST.get('email')
        password=request.POST.get('password')
        en=contactEnquiry(name=name,email=email,password=password)
        en.save()
        return redirect('home')
    return render(request,'contact.html')


@login_required(login_url='login')
def contactPage(request):
    
    return redirect('saveEnquiry')

@login_required(login_url='login')
def UpdateBlog(request):
    contactdata=contactEnquiry.objects.all()
    data={'contactdata':contactdata}
    return render(request,'update.html',data)

@login_required(login_url='login')
def EditBlog(request,id):

    contactdata=contactEnquiry.objects.get(id=id)

    return render(request,'edit.html',{'contactdata':contactdata})

@login_required(login_url='login')
def up(request,id):
    
    contactdata=contactEnquiry.objects.get(id=id)
    form=EmployeeForm(request.POST,instance=contactdata)
    if form.is_valid():
        form.save()
        return redirect('/update')

    return render(request,'edit.html',{'contactdata':contactdata})
@login_required(login_url='login')
def deleteData(request,id):
    
    contactdata=contactEnquiry.objects.get(id=id)
    contactdata.delete()
    
    return redirect('/update')

@login_required(login_url='login')
def searchEmp(request):
    
    contactdata=contactEnquiry.objects.all()

    filters=OrderFilters(request.GET,queryset=contactdata)

   
    
    return render(request,'searchEmployee.html',{'filters':filters})   
@login_required(login_url='login')
def searchBar(request):
    if request.method=='GET':
        query=request.GET.get('query')
        if query:
            data=contactEnquiry.objects.filter(name__icontains=query)
            return render(request,'home.html',{'data':data})
        else:
            print('No Information about it')
            return render(request,'home.html',{})
        
def is_valid_queryparam(param):
    return param != '' and param is not None


def countries_gdp_list(request):
    qs = CountryGDP.objects.order_by('name')

    name = request.GET.get('name')
    year = request.GET.get('year')

    request.session['name'] = name
    request.session['year'] = year

    if is_valid_queryparam(name):
        qs = qs.filter(name__icontains=name)

    if is_valid_queryparam(year):
        qs = qs.filter(year=year)

    page = request.GET.get('page', 1)
    paginator = Paginator(qs, 30)

    try:
        qs = paginator.page(page)
    except PageNotAnInteger:
        qs = paginator.page(1)
    except EmptyPage:
        qs = paginator.page(paginator.num_pages)

    context = {
        'countries_list': qs,
        'name': name,
        'year':year,
    }
    return render(request, "countries_list.html", context)




def countries_gdp_excel(request):

   
    
    qs = CountryGDP.objects.order_by('name')

    if 'name' in request.session:
        name = request.session['name']
    else:
        name = None

    if 'year' in request.session:
        year = request.session['year']
    else:
        year = None

    if is_valid_queryparam(name):
        qs = qs.filter(name__icontains=name)

    if is_valid_queryparam(year):
        qs = qs.filter(year=year)

    if year is None or year == '':
        year = "2019 - 2021"
    else:
        year = year

    if name is None or name == '':
        name = "All Countries"
    else:
        name = name

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'Countries GDP List' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:D1')
    worksheet.merge_cells('A2:D2')
    first_cell = worksheet['A1']
    first_cell.value = "Countries GDP List" + " " + year
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    second_cell = worksheet['A2']
    second_cell.value = name
    second_cell.font  = Font(bold=True, color="246ba1")
    second_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Countries GDP List' + " " + year

    # Define the titles for columns
    columns = ['Country Name','Country Code','Year', 'Value in USD']
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font  = Font(bold=True, color="F7F6FA")
        third_cell = worksheet['D3']
        third_cell.alignment = Alignment(horizontal="right")

    for countries in qs:
        row_num += 1

        # Define the data for each cell in the row
        row = [countries.name,countries.code,countries.year,countries.value]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
         cell = worksheet.cell(row=row_num, column=col_num)
         cell.value = cell_value
         if isinstance(cell_value, decimal.Decimal):
             cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

	        

    workbook.save(response)
    return response

def savecountries(request):
    if request.method=='POST':
        name=request.POST.get('name')
        code=request.POST.get('code')
        year=request.POST.get('year')
        value=request.POST.get('value')
        en=CountryGDP(name=name,code=code,value=value,year=year)
        en.save()
        return redirect('countries_gdp_list')
    return render(request,'addcountries.html')

def importExcel(request):

    if request.method=='POST':
       
        dataset=Dataset()
        new_person=request.FILES['my_file']
        imported_data=dataset.load(new_person.read(),format='xlsx')
        
        for data in imported_data:
            value=CountryGDP(data[0],
                             data[1],
                             data[2],
                             data[3],
                             data[4]
                             )
            value.save()
        return redirect('countries_gdp_list')

    return render(request,'form.html')

def gallery(request):


    if request.method == 'POST':
        form = ImageForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            return HttpResponse('successfully uploaded')
    else:
        form = ImageForm()
        img=Image.objects.all()

    return render(request, "galary.html", {'img':img,"form": form})


#cascading dropdown
def Cascade(request):
    return render(request, 'cascade.html')

def person_create_view(request):
    form = PersonCreationForm()
    if request.method == 'POST':
        form = PersonCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('person_add')
    return render(request, 'cascade.html', {'form': form})


def person_update_view(request, pk):
    person = get_object_or_404(Person, pk=pk)
    form = PersonCreationForm(instance=person)
    if request.method == 'POST':
        form = PersonCreationForm(request.POST, instance=person)
        if form.is_valid():
            form.save()
            return redirect('person_change', pk=pk)
    return render(request, 'cascade.html', {'form': form})


# AJAX
def load_cities(request):
    country_id = request.GET.get('country_id')
    cities = City.objects.filter(country_id=country_id).all()
    return render(request, 'city_dropdown_list_options.html', {'cities': cities})
    # return JsonResponse(list(cities.values('id', 'name')), safe=False)


def json(request):

   
    datas=Admin_add_employee.objects.all()

    with open(r'C:/Users/doid/Desktop/django/jsonfile/data.json', "w") as out:
        datas2=serializers.serialize('json',datas)
        out.write(datas2)

    
    data=list(Admin_add_employee.objects.values())
    return JsonResponse(data,safe=False)

def jsonITs(request):
    
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
         
         data=list(Admin_add_employee.objects.values())
         datas2=serializers.serialize('json',data)
         return JsonResponse(datas2,safe=True)

    return render(request,'jsonFile.html')

def dp(request):
    return render(request, 'dropdown.html')

def AdminEmployeeList(request):

    contactdata=Admin_add_employee.objects.all()
    
    filters=OrderFilters(request.GET,queryset=contactdata)

    data={'contactdata':contactdata,'filters':filters}

    return render (request,'Admin_EmployeeList.html',data)



# def AdminEmployeeList(request):
   
#     if request.method=="GET":

#         contactdata=Admin_add_employee.objects.all()
    
#         taskserializer=serializers.serialize("json",contactdata)

#         return JsonResponse(taskserializer,safe=False)
#     return JsonResponse({'message':'wrong validation'})


def AdminaddSaveData(request):
    if request.method=='POST':
        emp_id=request.POST.get('emp_id')
        emp_name=request.POST.get('emp_name')
        emp_post_name=request.POST.get('emp_post_name')

        emp_grade=request.POST.get('emp_grade')
        emp_salary=request.POST.get('emp_salary')
        emp_bonus=request.POST.get('emp_bonus')

        emp_increment=request.POST.get('emp_increment')
        emp_joining_date=request.POST.get('emp_joining_date')
        emp_joining_month=request.POST.get('emp_joining_month')


        emp_joining_year=request.POST.get('emp_joining_year')

        en=Admin_add_employee(emp_id=emp_id,emp_name=emp_name,emp_post_name=emp_post_name,emp_grade=emp_grade,
                          emp_salary=emp_salary,emp_bonus=emp_bonus,emp_increment=emp_increment,
                          emp_joining_date=emp_joining_date,emp_joining_month=emp_joining_month,
                          emp_joining_year=emp_joining_year)
        en.save()
        success='data saved successfully'+emp_id
        return HttpResponse(success)

        # return redirect('adminemployeeList')
    return render(request,'Add_emplyee.html')


def admin_update(request):

    contactdata=Admin_add_employee.objects.all()
    data={'contactdata':contactdata}
    return render(request,'Admin_Employee_update.html',data)

def adminemployeeEdit(request,id):

    contactdata=Admin_add_employee.objects.get(id=id)

    return render(request,'Edit_employee.html',{'contactdata':contactdata})

def adminup(request,id):

    


    # if request.method=="post":

    #     contactdata=Admin_add_employee.objects.get(id=id)
    #     send_data={"emp_id":contactdata.emp_id,"emp_name":contactdata.emp_name,"emp_post_name":contactdata.emp_post_name,
    #                "emp_grade":contactdata.emp_grade,"emp_salary":contactdata.emp_salary,"emp_bonus":contactdata.emp_bonus,
    #                "emp_increment":contactdata.emp_increment,"emp_joining_date":contactdata.emp_joining_date,"emp_joining_month":contactdata.emp_joining_month,
    #                "emp_joining_year":contactdata.emp_joining_year,}
    #     return JsonResponse(send_data)

    contactdata=Admin_add_employee.objects.get(id=id)
    form=AdminEmployeeForm(request.POST,instance=contactdata)
    if form.is_valid():
        
        form.save()

        return redirect('/adminemplist/adsming')

    return render(request,'Edit_employee.html',{'contactdata':contactdata})

# try edit Id with ajax
def ediId(request):

    if request.method=="POST":

        id=request.POST.get('id')
        contactdata=Admin_add_employee.objects.get(pk=id)

        form=AdminEmployeeForm(request.POST,instance=contactdata)
        if form.is_valid():
            form.save()
        data={"id":contactdata.id,"emp_id":contactdata.emp_id,
              "emp_name":contactdata.emp_name,"emp_post_name":contactdata.emp_post_name,
              "emp_grade":contactdata.emp_grade,"emp_salary":contactdata.emp_salary,
              "emp_bonus":contactdata.emp_bonus,"emp_increment":contactdata.emp_increment,
              "emp_joining_date":contactdata.emp_joining_date,"emp_joining_month":contactdata.emp_joining_month,
              "emp_joining_year":contactdata.emp_joining_year,
              }
        status='edit'
        print(status)
        return JsonResponse({'data':data,'status':status})

    #     return redirect('/adminemplist/adsming')

    return render(request,'Edit_employee.html',{'contactdata':contactdata})
#try delete ID with ajax
def deleteId(request):

    if request.method=="POST":

        id=request.POST.get('id')
        print(id)
        contactdata=Admin_add_employee.objects.get(pk=id)
        
        contactdata.delete()
        return JsonResponse({'status':1})
    else:
        return JsonResponse({'status':0})


#try deletemultiple ID with ajax
def deletmultipleId(request):

    if request.method=="POST":
        
        getidnew=request.POST.getlist('id[]')

        for getid in getidnew:
            
            contactdata=Admin_add_employee.objects.get(pk=getid)
            contactdata.delete()
        return JsonResponse({'status':1})
    else:
        return JsonResponse({'status':0})
        
    # return redirect('/adminemplist/adsming')


def admindeleteData(request,id):
    
    contactdata=Admin_add_employee.objects.get(id=id)
    contactdata.delete()
    
    return redirect('/adminemplist/adsming')

def adminMultipledeleteData(request,id=None):
    
    if request.method=="POST":

        getidnew=request.POST.getlist('mycheckbox')

        for getid in getidnew:
            
            contactdata=Admin_add_employee.objects.get(id=getid)
            contactdata.delete()


    # contactdata=Admin_add_employee.objects.get(id=id)
    # contactdata.delete()
    
    return redirect('/adminemplist/adsming')


def _valid_queryparam(param):
    return param != '' and param is not None

# def FilteradminEmployeeList(request):

#     qs = Admin_add_employee.objects.order_by('emp_name')

#     id = request.GET.get('id')
#     name = request.GET.get('name')
#     month = request.GET.get('month')

#     # request.session['id'] = id
#     # request.session['name'] = name
#     # request.session['month'] = month

#     if _valid_queryparam(id):
#         qs = Admin_add_employee.objects.filter(emp_id=id)
        

#     if _valid_queryparam(name):
#         qs = Admin_add_employee.objects.filter(emp_name__icontains=name)

#     if _valid_queryparam(month):
#         qs = Admin_add_employee.objects.filter(emp_joining_month=month)

#     page = request.GET.get('page', 1)
#     paginator = Paginator(qs, 30)

#     try:
#         qs = paginator.page(page)
#     except PageNotAnInteger:
#         qs = paginator.page(1)
#     except EmptyPage:
#         qs = paginator.page(paginator.num_pages)


#     # contactdata=Admin_add_employee.objects.all()


#     context = {
#         'EmployeeList': qs,
#         'id': id,
#         'name':name,
#         # 'contactdata':contactdata,
#     }
#     return render(request, "Admin_EmployeeList.html", context)



def FilteradminEmployeeList(request):

    ms=Admin_add_employee.objects.all()
    p=[]
    total_salary=[]
    count=[]
    status=[]
    id = request.POST.get('id')
    name = request.POST.get('name')
    month = request.POST.get('month')
    year = request.POST.get('year')
    context = {
        'EmployeeList':ms,
    }


    if request.method=="POST":
        
        if (id != '' and id is not None):
            qs1 = Admin_add_employee.objects.filter(emp_id=id).values()
            qs=list(qs1)
            p=qs
            status=id
            print("list of data",qs)
            
          
           
            
           
        
        elif (name != '' and name is not None):
            qs2 = Admin_add_employee.objects.filter(emp_name__icontains=name).values()
            qs=list(qs2)
            p=qs
            status=name
            print("list of data",qs)
        

        
        elif month and year:
            

            qs4e = Admin_add_employee.objects.filter(emp_joining_month=month,emp_joining_year=year).values()#return all list values

           

            qs=list(qs4e)
            p=qs
            status=3
            count=len(p)

            print("list of data",qs)

            
            total_salarys = sum(int(employee['emp_salary']) for employee in qs)

            
            total_salary=total_salarys

            print("total salary is",total_salary)
            
        else:
            return JsonResponse("No Information available!")
        
        

        return JsonResponse({'p':p,'received_salary':total_salary,'total_employee':count,'received_status':status})
    
    return render(request, 'Admin_EmployeeList.html', context)
        
   
   
  
def redata():

    contactdata=Admin_add_employee.objects.all()

    datas={
            'city':'Dhaka',
            'state':'Gazipur',
            'email':'rsmahbub@gmail.com',
            'address': "Dhaka",
            'phone':'01927722934',
            'contactdata':contactdata,
             
        }

    return datas

class ViewPDF(View):
	def get(self, request, *args, **kwargs):

		pdf = html_to_pdf('pdf_template.html', redata())
		return HttpResponse(pdf, content_type='application/pdf')
    




# class DownloadPDF(View):
# 	def get(self, request, *args, **kwargs):
      
# 		pdf = html_to_pdf('pdf_template.html', redata())

# 		response = HttpResponse(pdf, content_type='application/pdf')
# 		filename = "Invoice_%s.pdf" %("12341231")
# 		content = "attachment; filename='%s'" %(filename)
# 		response['Content-Disposition'] = content
# 		return response
    

def DownloadPDF(request):
	
      
    pdf = html_to_pdf('pdf_template.html', redata())

    response = HttpResponse(pdf, content_type='application/pdf')

    filename = "Invoice_%s.pdf" %("12341231")
    content = "attachment; filename='%s'" %(filename)
    response['Content-Disposition'] = content
    
    return response
    





   



# def downloadPDF(request):

#     contactdata=Admin_add_employee.objects.all()

#     response=HttpResponse(content_type='application/json')
#     filename = "Invoice_%s.pdf" %("12341231")
#     response['Content-Disposition']="attachment; filename='%s'" %(filename)
#     pdf=canvas.Canvas(response,pagesize=letter)
#     pdf.setTitle('PDF Report')
#     headers=["emp_id","emp_name","emp_post_name",
#              "emp_grade","emp_salary","emp_bonus",
#              "emp_increment","emp_joining_date","emp_joining_month"
#              ,"emp_joining_year"]
#     data=[headers]

#     for obj in contactdata:
#         data_row=[obj.emp_id,obj.emp_name,obj.emp_post_name,
#                       obj.emp_grade,obj.emp_salary,obj.emp_bonus,obj.emp_increment,
#                       obj.emp_joining_date,obj.emp_joining_month,obj.emp_joining_year]
#         data.append(data_row)

#     print(data)

#     table=Table(data)
#     table.setStyle(TableStyle([
#         ('BACKGROUND', (0,0), (-1,0),colors.green),
#         ('GRID', (0,1), (-1,-1), 1, colors.black)
#     ]))
#     canvas_width=600
#     canvas_height=600
#     table.wrapOn(pdf,canvas_width,canvas_height)
#     table.drawOn(pdf,1,canvas_height-len(data))
    
#     pdf.save()
#     return response
    
def reportLabdownloadPDF(request):

    contactdata=Admin_add_employee.objects.all()

    # qs = Admin_add_employee.objects.order_by('emp_name')


    my_path='D:\\ReportlabPractise\\myreport.pdf'


    headers=["emp_id","emp_name","emp_post_name",
             "emp_grade","emp_salary","emp_bonus",
             "emp_increment","emp_joining_date","emp_joining_month"
             ,"emp_joining_year"]
    data=[headers]

    for obj in contactdata:
        data_row=[obj.emp_id,obj.emp_name,obj.emp_post_name,
                      obj.emp_grade,obj.emp_salary,obj.emp_bonus,obj.emp_increment,
                      obj.emp_joining_date,obj.emp_joining_month,obj.emp_joining_year]
        data.append(data_row)

    print(data)

    # my_doc=SimpleDocTemplate(my_path,pagesize=(1024,900))
    t=Table(data,rowHeights=20,repeatRows=1)

    #others

    pdf=canvas.Canvas(my_path,pagesize=(1024,900))
   
    
    frameh=Frame(0,20,900,30,leftPadding=6, bottomPadding=6,
            rightPadding=6, topPadding=6, id=None,showBoundary=1)
    flow_obj1=[]
    tableh=Table([["Employee Information"]])
    flow_obj1.append(tableh)
    frameh.addFromList(flow_obj1,pdf)
    

    framef=Frame(0,40,900,30,showBoundary=1)
    flow_obj2=[]
    tablef=Table([["footer Employee Information"]])
    flow_obj2.append(tablef)
    framef.addFromList(flow_obj2,pdf)


    framefs=Frame(10,0,900,900,leftPadding=6, bottomPadding=6,
            rightPadding=6, topPadding=6, id=None,showBoundary=0)
    flow_obj3=[]
    t=Table(data,rowHeights=20,repeatRows=1)
    t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0),colors.green),
            ('GRID', (0,1), (-1,-1), 1, colors.black),
            ('FONTSIZE',(0,0),(-1,-1),10),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('BACKGROUND',(3,0),(-3,-1),colors.yellow)
        ]))
    flow_obj3.append(t)
    
    framefs.addFromList(flow_obj3,pdf)

    
    



    # pdf.drawString(10,10,"ok")



    pdf.save()
    

    #others





    # t.setStyle(TableStyle([
    #     ('BACKGROUND', (0,0), (-1,0),colors.green),
    #     ('GRID', (0,1), (-1,-1), 1, colors.black),
    #     ('FONTSIZE',(0,0),(-1,-1),10),
    #     ('ALIGN', (0,0), (-1,0), 'CENTER'),
    #     ('BACKGROUND',(3,0),(-2,-1),colors.chocolate)

    # ]))
    

    # elements=[]
    # elements.append(tableh)
    # elements.append(t)
   
    
    # my_doc.build(elements)
    

    
    
    return HttpResponse(content_type='application/json')

      
	#reportlab practise and save data as pdf	
# def generate_pdf(request):
   

#     my_path='D:\\ReportlabPractise\\myreport.pdf'

#     doc = SimpleDocTemplate(my_path, pagesize=(1024,900))

#     styles = getSampleStyleSheet()
#     flowables = []


#     custom_style = ParagraphStyle(
#     name='CustomStyle',
#     fontName='Helvetica',
#     fontSize=30,
#     leading=39,
#     textColor=colors.darkblue,
#     backColor=colors.whitesmoke,
    
#     spaceAfter=17,
    
#     rightIndent=20,
#     firstLineIndent=30,
#     alignment=1,  # 0=left, 1=center, 2=right, 4=justify
#     )
    
    
#     # Header
#     header_text = "Bangladesh IIIAC Industry Limited"
#     header = Paragraph(header_text, custom_style)
#     flowables.append(header)
#     flowables.append(Spacer(3, 16))

   

   

    

#     Bill=[["Bill To :","Mr. John Smith","","","","FROM :","DELHA SOURCE IIIT",""],
#           [" ","Vetani BBC","","","","","P.O OSTIA",""],
#           [" ","CANADA","","","","","CANADA 00099009",""],
#           ["","089 900989","","","","","TEL:7979798737",""],
#           ]
    
#     table2=Table(Bill,hAlign='LEFT',colWidths=100,rowHeights=20,rowSplitRange=10,spaceAfter=12)
#     table2.setStyle(TableStyle([
#         ('FONTSIZE', (0, 0), (-1, -1), 18),
      
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
       

#     ]))
#     flowables.append(table2)

#     flowables.append(Spacer(1, 20))
    

#     img_path1='C:\\Users\\doid\\Desktop\\QRCODE.jpg'
#     img3=Image(img_path1,width=100,height=100)

#     product=[["Product ID","Description","Qty","Unit Price","Line Total"],
#           ["P1003 ","Filling","10","420.00","4,200.00"],
#           ["P1000 ","ToothBrush","10","420.00","4,200.00"],
#           ["P1004 ","Root Canal","10","420.00","4,200.00"],
#           ["","","","Subtotal","13,417"],
#           ["","","PST","6.50%","807.5"],
#           ["","","GST","2.3%","345"],
#           ["","","","Shipping and handling",""],
#           ["","","","Total",""],
#           ["","","","Paid",""],
#           [img3,"","","Total Due",""],
#           ]
    
#     table3=Table(product,hAlign='LEFT',colWidths=160,rowHeights=20,rowSplitRange=10,spaceAfter=12)
#     for i in range(0, 4):
#         # Alternate row colors
#         if i % 2 == 0:
#             bg_color = colors.lightgrey
#         else:
#             bg_color = colors.whitesmoke
#         table3.setStyle(TableStyle([
#         ('FONTSIZE', (0, 0), (-1, -1), 15),
       
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

#         ('BACKGROUND', (0, i), (-1, -1), bg_color),

        


#         ('BOX', (0, 0), (-1, 3), 1, colors.black),
#         ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
#         ('BOX', (-1, 0), (-1, -1), 1, colors.black),

       
#         ('BOX', (1, 0), (1, 3), 1, colors.black),
#         ('BOX', (2, 0), (2, 3), 1, colors.black),
#         ('BOX', (3, 0), (3, 3), 1, colors.black),

#         ('GRID', (-1, 4), (-1, -1), 1, colors.black),
        
       

#         ]))
#     flowables.append(table3)





#     flowables.append(Spacer(1, 20))
   
    
#     # flowables.append(img1)
#     flowables.append(Spacer(1, 20))

#     img_path='C:\\Users\\doid\\Desktop\\sg.jpg'
#     img2=Image(img_path,width=300,height=300)
    
   
    
#     # flowables.append(img2)



#     flowables.append(Spacer(2, 15))
    

#     headers=["emp_id","emp_name","emp_post_name",
#              "emp_grade","emp_salary","emp_bonus",
#              "emp_increment","emp_joining_date","emp_joining_month"
#              ,"emp_joining_year"]
    

#      # Table
#     contactdata=Admin_add_employee.objects.all()
#     data = [headers]

#     for obj in contactdata:
#         data_row=[obj.emp_id,obj.emp_name,obj.emp_post_name,
#                       obj.emp_grade,obj.emp_salary,obj.emp_bonus,obj.emp_increment,
#                       obj.emp_joining_date,obj.emp_joining_month,obj.emp_joining_year]
#         data.append(data_row)

#     table = Table(data)
#     for i in range(1, len(contactdata)):
#         # Alternate row colors
#         if i % 2 == 0:
#             bg_color = colors.whitesmoke
#         else:
#             bg_color = colors.lightgrey
#         table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.orange),
#                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
#                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#                                ('BACKGROUND', (0, i), (-1, -1), bg_color),

                              

#                                ('GRID', (0, 0), (-1, -1), 1, colors.green)]))
#     flowables.append(table)




#     custom_style_footer = ParagraphStyle(
#     name='CustomStyle',
#     fontName='Helvetica',
#     fontSize=20,
#     leading=30,
#     textColor=colors.black,
#     backColor=colors.whitesmoke,
    
#     spaceAfter=17,
    
#     rightIndent=20,
#     firstLineIndent=30,
#     alignment=0,  # 0=left, 1=center, 2=right, 4=justify
#     )
#     # Footer
#     footer_text = "This is the footer"
#     footer = Paragraph(footer_text,custom_style_footer)
#     flowables.append(Spacer(1,20))
#     flowables.append(footer)

    

#     doc.build(flowables)
    
#     return HttpResponse(content_type='application/json')

def sgnpdf(request):
    data=[]

    status=request.POST.get('status')

    id=request.POST.get('id')
    
    name=request.POST.get('name')

    month=request.POST.get('month')

    year=request.POST.get('year')


    if(status==id):
       

        data=Admin_add_employee.objects.filter(emp_id=id)

    elif(status==name):

        
        data=Admin_add_employee.objects.filter(emp_name=name)
    else:
       
       data=Admin_add_employee.objects.filter(emp_joining_year=year,emp_joining_month=month)

    return data



def generate_pdf(request):


   
    y=sgnpdf(request)
       
    p=[]
   

    if request.method=='POST':

        if (y!=0 and y is not None):

            my_path='D:\\ReportlabPractise\\myreport.pdf'

            doc = SimpleDocTemplate(my_path, pagesize=(1024,900))

            styles = getSampleStyleSheet()
            flowables = []


            custom_style = ParagraphStyle(
            name='CustomStyle',
            fontName='Helvetica',
            fontSize=30,
            leading=39,
            textColor=colors.darkblue,
            backColor=colors.whitesmoke,
            
            spaceAfter=17,
            
            rightIndent=20,
            firstLineIndent=30,
            alignment=1,  # 0=left, 1=center, 2=right, 4=justify
            )
            
            
            # Header
            header_text = "Bangladesh IIIAC Industry Limited"
            header = Paragraph(header_text, custom_style)
            flowables.append(header)
            flowables.append(Spacer(3, 16))

        

        

            

            Bill=[["Bill To :","Mr. John Smith","","","","FROM :","DELHA SOURCE IIIT",""],
                [" ","Vetani BBC","","","","","P.O OSTIA",""],
                [" ","CANADA","","","","","CANADA 00099009",""],
                ["","089 900989","","","","","TEL:7979798737",""],
                ]
            
            table2=Table(Bill,hAlign='LEFT',colWidths=100,rowHeights=20,rowSplitRange=10,spaceAfter=12)
            table2.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 18),
            
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            

            ]))
            flowables.append(table2)

            flowables.append(Spacer(1, 20))
            

            img_path1='C:\\Users\\doid\\Desktop\\QRCODE.jpg'
            img3=Image(img_path1,width=100,height=100)

            product=[["Product ID","Description","Qty","Unit Price","Line Total"],
                ["P1003 ","Filling","10","420.00","4,200.00"],
                ["P1000 ","ToothBrush","10","420.00","4,200.00"],
                ["P1004 ","Root Canal","10","420.00","4,200.00"],
                ["","","","Subtotal","13,417"],
                ["","","PST","6.50%","807.5"],
                ["","","GST","2.3%","345"],
                ["","","","Shipping and handling",""],
                ["","","","Total",""],
                ["","","","Paid",""],
                [img3,"","","Total Due",""],
                ]
            
            table3=Table(product,hAlign='LEFT',colWidths=160,rowHeights=20,rowSplitRange=10,spaceAfter=12)
            for i in range(0, 4):
                # Alternate row colors
                if i % 2 == 0:
                    bg_color = colors.lightgrey
                else:
                    bg_color = colors.whitesmoke
                table3.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 15),
            
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

                ('BACKGROUND', (0, i), (-1, -1), bg_color),

                


                ('BOX', (0, 0), (-1, 3), 1, colors.black),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
                ('BOX', (-1, 0), (-1, -1), 1, colors.black),

            
                ('BOX', (1, 0), (1, 3), 1, colors.black),
                ('BOX', (2, 0), (2, 3), 1, colors.black),
                ('BOX', (3, 0), (3, 3), 1, colors.black),

                ('GRID', (-1, 4), (-1, -1), 1, colors.black),
                
            

                ]))
            flowables.append(table3)





            flowables.append(Spacer(1, 20))
        
            
            # flowables.append(img1)
            flowables.append(Spacer(1, 20))

            img_path='C:\\Users\\doid\\Desktop\\sg.jpg'
            img2=Image(img_path,width=300,height=300)
            
        
            
            # flowables.append(img2)



            flowables.append(Spacer(2, 15))
            

            headers=["emp_id","emp_name","emp_post_name",
             "emp_grade","emp_salary","emp_bonus",
             "emp_increment","emp_joining_date","emp_joining_month"
             ,"emp_joining_year"]
            

            # Table
           
            rs=y
            data = [headers]

            for obj in rs:
                data_row=[obj.emp_id,obj.emp_name,obj.emp_post_name,
                       obj.emp_grade,obj.emp_salary,obj.emp_bonus,obj.emp_increment,
                       obj.emp_joining_date,obj.emp_joining_month,obj.emp_joining_year]
                data.append(data_row)

            table = Table(data)
            table.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOX',(0, 0), (-1, -1),1,colors.black),
                ('GRID',(0, 0), (-1, -1),1,colors.black),
            

            ]))

        
            flowables.append(table)




            custom_style_footer = ParagraphStyle(
            name='CustomStyle',
            fontName='Helvetica',
            fontSize=20,
            leading=30,
            textColor=colors.black,
            backColor=colors.whitesmoke,
            
            spaceAfter=17,
            
            rightIndent=20,
            firstLineIndent=30,
            alignment=0,  # 0=left, 1=center, 2=right, 4=justify
            )
            # Footer
            footer_text = "This is the footer"
            footer = Paragraph(footer_text,custom_style_footer)
            flowables.append(Spacer(1,20))
            flowables.append(footer)

            

            doc.build(flowables)
    
        
   
    return HttpResponse(content_type='application/json')






# def generate_pdf_with_watermark(request):
#     # Create a response object
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="invoice.pdf"'
#     doc = SimpleDocTemplate(response, pagesize=letter)

#     # Invoice details
#     invoice_number = "INV-001"
#     invoice_date = "May 10, 2024"
#     company_name = "Your Company Name"
#     company_address = "123 Street, City, Country"
#     client_name = "Client Name"
#     client_address = "456 Street, City, Country"

#     # Item list
#     items = [
#         {"item": "Item 1", "quantity": 2, "unit_price": 50, "total": 100},
#         {"item": "Item 2", "quantity": 1, "unit_price": 75, "total": 75},
#     ]

#     # Create a list of flowables
#     elements = []

#     # Add company logo
#     logo_path = "C:\\Users\\doid\\Desktop\\sg.jpg"  # Path to your company logo
#     logo = Image(logo_path, width=500, height=50)
#     elements.append(logo)

#     # Add invoice details
#     styles = getSampleStyleSheet()
#     elements.append(Paragraph(f"Invoice Number: {invoice_number}", styles["Heading1"]))
#     elements.append(Paragraph(f"Invoice Date: {invoice_date}", styles["Normal"]))
#     elements.append(Paragraph(f"From: {company_name}<br/>{company_address}", styles["Normal"]))
#     elements.append(Paragraph(f"To: {client_name}<br/>{client_address}", styles["Normal"]))

#      # Add paragraphs at specific positions
   
#     width = 300
#     height = 100
#     p3 = Paragraph("Paragraph 3", styles["Normal"])
   
    

#     elements.append(p3)


#     # Add item list
#     data = [["Item", "Quantity", "Unit Price", "Total"]]
#     for item in items:
#         data.append([item["item"], item["quantity"], item["unit_price"], item["total"]])
#     table = Table(data)
#     table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#                                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
#                                ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
#     elements.append(table)

#     # Add subtotal, taxes, and total
#     subtotal = sum(item["total"] for item in items)
#     elements.append(Paragraph(f"Subtotal: ${subtotal}", styles["Normal"]))
#     elements.append(Spacer(1, 12))

#     # Build the PDF document

#     doc.build(elements)

#     return response

import phonenumbers
from phonenumbers import geocoder
from phonenumbers import carrier
import opencage
from opencage.geocoder import OpenCageGeocode


def generate_pdf_with_watermark(request):
    numbers=input('Enter your number')
    new_number=phonenumbers.parse(numbers)
    location=geocoder.description_for_number(new_number,'en')
    print(location)

    service_name=carrier.name_for_number(new_number,'en')
    print(service_name)
    key=["03fbe7b932e3418f9b770cf3a35ed4cd"]
    geocoders=OpenCageGeocode(key)
    query=str(location)
    result=geocoders.geocode(query)
    print(result)

    lat=result[0]['geometry']['lat']
    lng=result[0]['geometry']['lng']

    print(lat,lng)


    return  redirect('/adminemplist/adsming')

#https://www.youtube.com/watch?v=yEcYSNL9rx0


def funcpdf(request):
    return render(request,'pdfresult.html')


def AdminEmployeeexcel(request):

    contactdata=Admin_add_employee.objects.all()


    datas={
            'city':'Dhaka',
            'state':'Gazipur',
            'email':'rsmahbub@gmail.com',
            'address': "Dhaka",
            'phone':'01927722934',
            
            
             
        }
    
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'employee' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:D1')
    worksheet.merge_cells('A2:D2')

    first_cell = worksheet['A1']
    first_cell.value = "Address List is " + " " + datas['address']
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    second_cell = worksheet['A2']
    second_cell.value ="Phone Number is "+"" + datas['phone']
    second_cell.font  = Font(bold=True, color="246ba1")
    second_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'phone List' + " " + datas['phone']

    # Define the titles for columns
    columns = ['Employee Id','Employee Name','Bonus','Salary']
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font  = Font(bold=True, color="F7F6FA")
        third_cell = worksheet['D3']
        third_cell.alignment = Alignment(horizontal="right")

    for countries in contactdata:
        row_num += 1

        # Define the data for each cell in the row
        # row = [datas['address'],datas['phone'],datas['city'],datas['email']]

        row = [countries.emp_id,countries.emp_name,countries.emp_bonus,countries.emp_salary]


        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
         cell = worksheet.cell(row=row_num, column=col_num)
         cell.value = cell_value
         if isinstance(cell_value, decimal.Decimal):
             cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

	        

    workbook.save(response)
    return response